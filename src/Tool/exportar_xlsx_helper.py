"""Gravacao de planilhas XLSX a partir de cabecalhos e linhas."""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from src.Tool.registrador_log import RegistradorLog

_CARACTERES_INVALIDOS_ARQUIVO = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def sanitizar_nome_arquivo(texto: str, *, padrao: str = "exportacao") -> str:
    """Remove caracteres invalidos para nome de arquivo no Windows."""
    limpo = (texto or "").strip()
    if not limpo:
        limpo = padrao
    limpo = _CARACTERES_INVALIDOS_ARQUIVO.sub("_", limpo)
    limpo = re.sub(r"\s+", "_", limpo)
    return limpo[:120] or padrao


def nome_arquivo_xlsx_com_data(nome_base: str) -> str:
    """Sugere nome com data/hora para evitar sobrescrever exportacoes."""
    base = sanitizar_nome_arquivo(nome_base)
    carimbo = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_{carimbo}.xlsx"


def salvar_planilha_xlsx(
    caminho: str | Path,
    cabecalhos: list[str],
    linhas: list[list[str]],
    *,
    nome_aba: str = "Dados",
) -> tuple[bool, str | None]:
    """Persiste cabecalhos e linhas em um arquivo XLSX."""
    if not cabecalhos:
        return False, "Nenhuma coluna para exportar."
    if not linhas:
        return False, "Nenhuma linha para exportar."

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False, "Biblioteca openpyxl nao instalada. Execute: pip install openpyxl"

    log = RegistradorLog()
    destino = Path(caminho)
    if destino.suffix.lower() != ".xlsx":
        destino = destino.with_suffix(".xlsx")

    try:
        destino.parent.mkdir(parents=True, exist_ok=True)
        planilha = Workbook()
        aba = planilha.active
        aba.title = sanitizar_nome_arquivo(nome_aba, padrao="Dados")[:31]

        aba.append(cabecalhos)
        for celula in aba[1]:
            celula.font = Font(bold=True)

        for linha in linhas:
            aba.append(linha)

        for indice, cabecalho in enumerate(cabecalhos, start=1):
            largura = max(len(str(cabecalho)), 10)
            for linha in linhas:
                if indice - 1 < len(linha):
                    largura = max(largura, len(str(linha[indice - 1])))
            aba.column_dimensions[get_column_letter(indice)].width = min(largura + 2, 48)

        planilha.save(destino)
        log.info(f"Planilha exportada: {destino.name}")
        return True, None
    except OSError as exc:
        log.erro(f"Falha ao salvar XLSX: {exc}")
        return False, f"Nao foi possivel salvar o arquivo: {exc}"
    except Exception as exc:
        log.erro(f"Falha ao exportar XLSX: {exc}")
        return False, f"Erro ao gerar planilha: {exc}"
